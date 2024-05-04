import os
import openpyxl

import Workspace
from Printermanager import Printermanager
import copy
import pickle
import platform
import ast
from User import User

class Outputmanager:

    def __init__(self, printermanager: Printermanager):
        self.printermanager = printermanager

    def return_deep_copy_of_printermanger_printers(self):
        return copy.deepcopy(self.printermanager.printers) #make a deep-copy of the list. this makes sure that we don't change the orignal list

    def create_output_excel_list_for_robot(self, path_with_filename: str, title_of_worksheet: str, list_with_header_names: list, printer_list: list):
        file_path = f'{path_with_filename}.xlsx'

        # check if the file exists and if it does delete it
        if os.path.exists(file_path):
            os.remove(file_path)

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = title_of_worksheet

        # get the instance variable names and write them as headers
        for i, variable in enumerate(list_with_header_names):
            worksheet.cell(row=1, column=i + 1, value=variable)

        # create a list of dictionarys of printername and papersources {printername: pstva1234, M0: papersourceobj1, S1: papersourceobj2}
        list_of_dict_with_printername_and_papersources = []

        #iterate through all printers in printermanager.printers
        for i, instance in enumerate(printer_list):
            list_of_dict = []

            papersources = getattr(instance, "papersources")

            for papersource in papersources:
                copies_papersource_for_each_wcps = []
                num_of_wcps = len(papersource.wcps)  # 0 = no wcps
                if num_of_wcps > 0:
                    for i in range(num_of_wcps): #generate as many copies of new dict as wcps in papersource
                        new_dict = {}   #generate a new_dict for each papersource
                        new_dict["Schacht Name"] = papersource.printerslot
                        new_dict["Druckername Printerserver"] = f"{instance.printername}-{papersource.printerslot}"
                        new_dict["Format des Formulars"] = papersource.paperformat
                        new_dict["Aktiv"] = papersource.active

                        if papersource.inspect == True:
                            new_dict["Inspect"] = "x"
                        elif papersource.inspect == "CUT":
                            new_dict["Inspect"] = "CUT"
                        else:
                            new_dict["Inspect"] = ""

                        if papersource.twosided == True:
                            if instance.driver == "Brother HL-L6250DN series":
                                new_dict["2-sided"] = "2-sided"
                            elif instance.driver == "Canon Generic Plus PCL6":
                                new_dict["2-sided"] = "2-sided Printing"
                            elif instance.driver == "Xerox VersaLink C9000":
                                new_dict["2-sided"] = "2-sided"
                            else:
                                raise Exception("Unkown Driver")
                        else:
                            if instance.driver == "Brother HL-L6250DN series":
                                new_dict["2-sided"] = "None"
                            elif instance.driver == "Canon Generic Plus PCL6":
                                new_dict["2-sided"] = "1-sided Printing"
                            elif instance.driver == "Xerox VersaLink C9000":
                                new_dict["2-sided"] = "None"
                            else:
                                raise Exception("Unkown Driver")




                        copies_papersource_for_each_wcps.append(new_dict) #now we have for each wcps a copy of the papersource dict with all the same values



                    #loop through wcps of papersource. wcps contains a tuple (workspace_name, cari-doc, printername, printerslot, department)
                    temp_wcps_list = [] # now we have a list of tuples of wcps which have to be added to copies_papersource
                    for wcps in papersource.wcps:
                        temp_wcps_list.append(wcps)

                    for i, dict in enumerate(copies_papersource_for_each_wcps):
                        dict["Formular CARI / Prüfbahn / Parkplatz"] = temp_wcps_list[i][1]  #tuple(1) = caridoc
                        dict["Zuständige Fachabteilung"] = temp_wcps_list[i][4]              #tuple(4) = deparment
                        dict["Arbeitsplatz (Büro)"] = temp_wcps_list[i][0]                   #tuple(0) = workspace_name

                    for item in copies_papersource_for_each_wcps:
                        list_of_dict.append(item)
                else:
                    new_dict = {}  # generate a new_dict for each papersource
                    new_dict["Schacht Name"] = papersource.printerslot
                    new_dict["Druckername Printerserver"] = f"{instance.printername}-{papersource.printerslot}"
                    new_dict["Format des Formulars"] = papersource.paperformat
                    new_dict["Aktiv"] = papersource.active

                    if papersource.inspect == True:
                        new_dict["Inspect"] = "x"
                    elif papersource.inspect == "CUT":
                        new_dict["Inspect"] = "CUT"
                    else:
                        new_dict["Inspect"] = ""

                    if papersource.twosided == True:
                        if instance.driver == "Brother HL-L6250DN series":
                            new_dict["2-sided"] = "2-sided"
                        elif instance.driver == "Canon Generic Plus PCL6":
                            new_dict["2-sided"] = "2-sided Printing"
                    elif papersource.twosided == False:
                        if instance.driver == "Brother HL-L6250DN series":
                            new_dict["2-sided"] = "None"
                        elif instance.driver == "Canon Generic Plus PCL6":
                            new_dict["2-sided"] = "1-sided Printing"
                    else:
                        new_dict["2-sided"] = ""

                    list_of_dict.append(new_dict)

            for dict in list_of_dict:
                dict["Standort"] = instance.standort
                dict["Bemerkung"] = instance.buero
                dict["Printserver Link"] = "\\\\vstvacp100001\\"
                dict["Druckername"] = instance.printername
                dict["IP Drucker (nur zur Information für Vergleich QIP)"] = instance.ip
                dict["Portname"] = instance.printername
                dict["Drucker Modell"] = instance.model
                dict["Drucker Treiber"] = instance.driver

                list_of_dict_with_printername_and_papersources.append(dict)

        # write the instance variable values as rows
        for i, dictionary in enumerate(list_of_dict_with_printername_and_papersources):
            value_list = []
            for variable in list_with_header_names:
                value_list.append(dictionary.get(variable))
            for j, value in enumerate(value_list):
                worksheet.cell(row=i + 2, column=j + 1, value=value)
        workbook.save(file_path)

    def create_output_excel_list_for_serdar_new_version(self, path_with_filename: str, title_of_worksheet: str, list_with_header_names: list, printer_list: list, delete_previous_file = True):
        file_path = f'{path_with_filename}.xlsx'

        # check if the file exists and if it does delete it
        if delete_previous_file:
            if os.path.exists(file_path):
                os.remove(file_path)
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = title_of_worksheet
        else:
            workbook = openpyxl.load_workbook(file_path)

        if title_of_worksheet in workbook.sheetnames:  # check if the worksheets exists
            worksheet = workbook[title_of_worksheet]  # access the existing worksheet
        else:
            worksheet = workbook.create_sheet(title_of_worksheet)  # create a new worksheet

        # get the instance variable names and write them as headers
        # list_with_header_names = ["printername", "paperslots", "users", "users_to_windowsprinter", "users_combined", "pcs_to_default_windowsprinter"]
        for i, variable in enumerate(list_with_header_names):
            worksheet.cell(row=1, column=i + 1, value=variable)

        #add printername in column 1
        for i, printer in enumerate(printer_list):
            # add printername in column 1
            worksheet.cell(row=i + 2, column=1, value = printer.printername)
            # add paperslots in column 2
            if (hasattr(printer, "printerslots_for_pickle")):
                value_printerslots = str(printer.printerslots_for_pickle)
                parsed_set = ast.literal_eval(value_printerslots)
                value_printerslots = ','.join(parsed_set)
            else:
                value_printerslots = ""
            worksheet.cell(row=i + 2, column=2, value= value_printerslots)
            # add users_for_cari in column 3
            value_cari = str(printer.user_to_windowsprinter_for_cari)
            parsed_set = ast.literal_eval(value_cari)
            value_cari = ','.join(parsed_set)
            worksheet.cell(row=i + 2, column=3, value=value_cari)
            # add users_windowsprinters in column 4
            value_user = str(printer.user_to_windowsprinter)
            parsed_set = ast.literal_eval(value_user)
            value_user = ','.join(parsed_set)
            worksheet.cell(row=i + 2, column=4, value=value_user)
            # add users_combined in column 5
            value = value_cari + "," + value_user
            value = "" if value == "," else value
            worksheet.cell(row=i + 2, column=5, value=value)
            # add pc to windowsprinter in 6
            value = str(printer.pc_to_default_windowsprinter)
            parsed_set = ast.literal_eval(value)
            value = ','.join(parsed_set)
            worksheet.cell(row=i + 2, column=6, value=value)

        workbook.save(file_path)

    def create_output_excel_list_for_serdar(self, path_with_filename: str, title_of_worksheet: str, list_with_header_names: list, printer_list: list, delete_previous_file = True):
        file_path = f'{path_with_filename}.xlsx'

        # check if the file exists and if it does delete it
        if delete_previous_file:
            if os.path.exists(file_path):
                os.remove(file_path)
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = title_of_worksheet
        else:
            workbook = openpyxl.load_workbook(file_path)

        if title_of_worksheet in workbook.sheetnames:  # check if the worksheets exists
            worksheet = workbook[title_of_worksheet]  # access the existing worksheet
        else:
            worksheet = workbook.create_sheet(title_of_worksheet)  # create a new worksheet


        # get the instance variable names and write them as headers
        # list_with_header_names = ["printername", "paperslots", "users", "users_to_windowsprinter", "users_combined", "pcs_to_default_windowsprinter"]
        for i, variable in enumerate(list_with_header_names):
            worksheet.cell(row=1, column=i + 1, value=variable)

        # printer_list = {printername = pstva1769, paperslots = [s1, s2, s3], workspace = [AL-ZUL-PEZ1,
        # AL_ZUL-PEZ2...], users = [B126SMP, B126IMD...], users_to_windowsprinter = [B126KEC, ...],
        # pcs_to_default_windowsprinter = [CSTVA1234, ...]}
        for i, dictionary in enumerate(printer_list):
            #dictionary is one printer like {{'printername': 'PSTVA3170', 'paperslots': ['S1', 'S2'], 'workspaces': [], 'users': [], 'users_to_windowsprinter': ['B126VOK'], 'pcs_to_default_windowsprinter': []}}
            value_list = []
            for variable in list_with_header_names:
                value_list.append(dictionary.get(variable))
            for j, value in enumerate(value_list):
                if (isinstance(value, list)):
                    value = ','.join(value)
                worksheet.cell(row=i + 2, column=j + 1, value=value)
        workbook.save(file_path)


    def create_output_excel_statistic(self, path_with_filename: str, title_of_worksheet: str, list_with_header_names: list, printer_list: list, delete_previous_file, workspace_list: Workspace):
        file_path = f'{path_with_filename}.xlsx'

        # check if the file exists and if it does delete it
        if delete_previous_file:
            if os.path.exists(file_path):
                os.remove(file_path)
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = title_of_worksheet
        else:
            workbook = openpyxl.load_workbook(file_path)

        if title_of_worksheet in workbook.sheetnames:  # check if the worksheets exists
            worksheet = workbook[title_of_worksheet]  # access the existing worksheet
        else:
            worksheet = workbook.create_sheet(title_of_worksheet)  # create a new worksheet



        list_of_all_users = set()

        for printer in printer_list:
            for user in printer["users"]:
                list_of_all_users.add(user)

        list_of_userobjects = []
        for username in list_of_all_users:
            list_of_userobjects.append(User(username))

        for workspace in workspace_list:
            for user in workspace.users:
                for userobj in list_of_userobjects:
                    if userobj.username == user:
                        userobj.count_cari_workspace()

        for printer in printer_list:
            for user in printer["users"]:
                for userobj in list_of_userobjects:
                    if userobj.username == user:
                        userobj.count_printer()
                        for paperslot in printer["paperslots"]:
                            userobj.count_printerslot()
        #we have now a list of users in userobjects which have the numbers of workspaces, the numbers of printers and the numbers of paperslots

        # get the instance variable names and write them as headers
        # list_with_header_names = ["username", "number_cari_workspaces", "number_printers", "number_printerslots"]
        for i, variable in enumerate(list_with_header_names):
            worksheet.cell(row=1, column=i + 1, value=variable)

        for row, obj in enumerate(list_of_userobjects, start=2):
            for col, header in enumerate(list_with_header_names, start=1):
                value = getattr(obj, header)
                worksheet.cell(row=row, column=col, value=value)


        workbook.save(file_path)

    def create_output_excel_list_for_gilles(self, path_with_filename: str, title_of_worksheet: str, list_with_header_names: list, printermanager: Printermanager):
        file_path = f'{path_with_filename}.xlsx'

        #if no excelfile exists, create a new excel file, otherwise load the existing file
        if not os.path.exists(f'{path_with_filename}.xlsx'):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = title_of_worksheet
        else:
            workbook = openpyxl.load_workbook(file_path)

        if title_of_worksheet in workbook.sheetnames: #check if the worksheets exists
            worksheet = workbook[title_of_worksheet] # access the existing worksheet
        else:
            worksheet = workbook.create_sheet(title_of_worksheet) # create a new worksheet

        mapping_location_to_id = {
            "AMA" : 1,
            "Albisgüetli" : 1,
            "Winterthur" : 2,
            "Regensdorf" : 3,
            "Hinwil" : 4,
            "Oberrieden" : 5,
            "Bülach" : 6,
            "Bassersdorf" : 7,
            "Tiefenbrunnen" : 8
        }

        mapping_locationid_to_lieugestion = {
            1 : "Albisgüetli",
            2 : "Winterthur",
            3 : "Regensdorf",
            4 : "Hinwil",
            5 : "Oberrieden",
            6 : "Bülach",
            7 : "Bassersdorf",
            8 : "Tiefenbrunnen"
        }

        # get the instance variable names and write them as headers
        for i, variable in enumerate(list_with_header_names):
            worksheet.cell(row=1, column=i + 1, value=variable)

        #after writing the header, set row = 2, which is directly after the header
        row = 2

        for workspace in printermanager.workspaces:
            # creates content for worksheet bureau
            if title_of_worksheet == "Bureau":
                id = workspace.id
                libelle = workspace.name
                value = workspace.name
                worksheet.cell(row=row, column=1, value=id)
                worksheet.cell(row=row, column=2, value=libelle)
                worksheet.cell(row=row, column=3, value=value)
                row += 1 # after writing the row increase the value that in the next iteration of the for loop we write the next row

            # creates content for worksheet LieuGestion
            elif title_of_worksheet == "LieuGestion":
                pass

            # creates content for worksheet LienBureauLieuGestion
            elif title_of_worksheet == "LienBureauLieuGestion":
                bureau = workspace.name
                _location = workspace.location
                lieuGestion = mapping_location_to_id[_location]
                worksheet.cell(row=row, column=1, value=bureau)
                worksheet.cell(row=row, column=2, value=lieuGestion)
                row += 1 # after writing the row increase the value that in the next iteration of the for loop we write the next row

            # creates content for worksheet BureauUsers
            elif title_of_worksheet == "BureauUsers":
                bureau_id = workspace.id
                bureau_libelle = workspace.name
                users = str(workspace.users).lower() #write username in lowercase .lower()
                worksheet.cell(row=row, column=1, value=bureau_id)
                worksheet.cell(row=row, column=2, value=bureau_libelle)
                worksheet.cell(row=row, column=3, value=users)
                row += 1  # after writing the row increase the value that in the next iteration of the for loop we write the next row

            else:
                raise Exception(f"{title_of_worksheet} cannot be used in this function")

        if title_of_worksheet == "LieuGestion":
            for key in mapping_locationid_to_lieugestion:
                id = key
                libelle = mapping_locationid_to_lieugestion[key]
                value = key
                worksheet.cell(row=row, column=1, value=id)
                worksheet.cell(row=row, column=2, value=libelle)
                worksheet.cell(row=row, column=3, value=value)
                row += 1  # after writing the row increase the value that in the next iteration of the for loop we write the next row

        workbook.save(file_path)

    def delete_printer_without_wcps(self, printer_list: list) -> list:
        """this function deletes all printerslots of a printer with no wcps. it means it is a printer used with
        cari. if a printer has not at least one printerslot with a wcps the whole printer is deleted"""
        temp_printers = []
        for printer in printer_list:
            temp_papersources = []
            for papersource in printer.papersources:
                #if papersource has no wcps objects do nothing
                if len(papersource.wcps) == 0:
                    continue
                #otherwise add the papersource to the temp_papersources_list
                else:
                    temp_papersources.append(papersource)

            #if temp_papersources has no items do nothing. it means there were no wcps present in papersource
            if temp_papersources == 0:
                continue
            #otherwise replace the papersources of the printer only with the papersources with a wcps
            else:
                printer.papersources = temp_papersources

        # now we ended up with printers without any papersources because for some printers there was not a single papersource with a wcps
        # we loop through the printers again and add all the printers to the temp_printers list which have at least one papersource (with a wcps)

        for printer in printer_list:
            # if the printer has no papersource left we dont add it to temp_printers
            if len(printer.papersources) == 0:
                continue
            # otherwise we add it
            else:
                temp_printers.append(printer)

        # no we switch self.printers with temp_printers
        printer_list = temp_printers

        return printer_list

    def delete_all_wcps(self, printer_list: list) -> list:
        """deletes all wcps of printers. papersource will remain. it means we end up with a printerlist without
        caridocs and workspaces attached"""
        for printer in printer_list:
            for papersource in printer.papersources:
                papersource.wcps = set()

        return printer_list

    def delete_papersource_with_inspect(self, printer_list: list) -> list:
        """deletes all papersource with inspect == true. if printer has no more papersource the whole printer gets deleted"""
        temp_printers = []
        for printer in printer_list:
            temp_papersources = []
            for papersource in printer.papersources:
                # if papersource has inspect == true do nothing
                if papersource.inspect == True or papersource.inspect == "CUT":
                    continue
                # otherwise add the papersource to the temp_papersources_list
                else:
                    temp_papersources.append(papersource)

            # if temp_papersources has no items do nothing. it means there were no papersources without inspect == true
            if temp_papersources == 0:
                continue
            # otherwise replace the papersources of the printer only with the papersources with inpsect == false
            else:
                printer.papersources = temp_papersources

        for printer in printer_list:
            # if the printer has no papersource left we dont add it to temp_printers
            if len(printer.papersources) == 0:
                continue
            # otherwise we add it
            else:
                temp_printers.append(printer)

        # no we switch self.printers with temp_printers
        printer_list = temp_printers

        return printer_list

    def pickle_save(self, list_to_save, output_folder: str, filename: str):

        os_name = platform.system()

        if os_name == "Windows":
            parent_dir = "output\pickle-files"
        else:
            parent_dir = "/output/pickle-files"

        path = os.path.join(parent_dir, output_folder)

        #create the outputfolder if it doesn't exist'
        os.makedirs(path, exist_ok=True)

        file_path = os.path.join(path, f"{filename}.pkl")
        print(file_path)
        #Open a file in binary write mode
        with open(file_path, 'wb') as file:
            pickle.dump(list_to_save, file)

    def pickle_load(self, input_folder: str, filename: str):

        os_name = platform.system()

        if os_name == "Windows":
            parent_dir = "output\pickle-files"
        else:
            parent_dir = "/output/pickle-files"

        path = os.path.join(parent_dir, input_folder)

        file_path = os.path.join(path, f"{filename}.pkl")

        with open(file_path, 'rb') as file:
            loaded_data = pickle.load(file)

        return loaded_data

    def return_delta_of_two_lists(self, list1: list, list2: list) -> list:
        """warning, method works only with simple list"""
        delta = [item for item in list1 if item not in list2] + [item for item in list2 if item not in list1]
        return delta
















